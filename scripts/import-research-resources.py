#!/usr/bin/env python3
"""Normalize the research field/society/journal workbook into app JSON masters."""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path

import openpyxl


def text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def number_id(value: object, prefix: str, fallback: int) -> str:
    raw = text(value)
    try:
        number = int(float(raw))
    except ValueError:
        number = fallback
    return f"{prefix}-{number:05d}"


def split_list(value: object, *, commas: bool = False) -> list[str]:
    raw = text(value)
    if not raw:
        return []
    pattern = r"\s*(?:／|\n|；|;)\s*"
    if commas:
        pattern = r"\s*(?:／|\n|；|;|、|，|,)\s*"
    return list(dict.fromkeys(part.strip() for part in re.split(pattern, raw) if part.strip()))


def status(url_type: str) -> str:
    return "official" if "公式" in url_type else "unverified"


def rows(sheet):
    return sheet.iter_rows(min_row=2, values_only=True)


def value(row: tuple[object, ...], index: int) -> str:
    return text(row[index]) if index < len(row) else ""


def write_json(path: Path, payload: object) -> None:
    path.write_text(json.dumps(payload, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")


def main() -> None:
    if len(sys.argv) < 3:
        raise SystemExit("usage: import-research-resources.py INPUT.xlsx OUTPUT_DIR")

    source = Path(sys.argv[1])
    output_dir = Path(sys.argv[2])
    output_dir.mkdir(parents=True, exist_ok=True)
    workbook = openpyxl.load_workbook(source, read_only=True, data_only=True)

    field_sheet = workbook["1_研究領域から探す"]
    fields = []
    for row_number, row in enumerate(rows(field_sheet), start=1):
        if not value(row, 0):
            continue
        hierarchy = [value(row, i) for i in range(1, 7)]
        name_ja = value(row, 7) or next((part for part in reversed(hierarchy) if part), "")
        if not name_ja:
            continue
        path_parts = list(dict.fromkeys(part for part in [*hierarchy, name_ja] if part))
        fields.append({
            "id": number_id(row[0], "field", row_number),
            "sourceNo": value(row, 0),
            "sourceSheet": field_sheet.title,
            "nameJa": name_ja,
            "nameEn": value(row, 8),
            "kingdom": hierarchy[0],
            "division": hierarchy[1],
            "className": hierarchy[2],
            "orderName": hierarchy[3],
            "family": hierarchy[4],
            "species": hierarchy[5],
            "level": value(row, 9),
            "definition": value(row, 20) or value(row, 10),
            "beginnerDescription": value(row, 19),
            "researchObjects": split_list(value(row, 21), commas=True),
            "methods": split_list(value(row, 22), commas=True),
            "researchPurpose": value(row, 23),
            "representativeThemes": split_list(value(row, 24), commas=True),
            "adjacentDifference": value(row, 25),
            "evidenceSources": [item for item in [
                {"url": value(row, 26), "type": value(row, 27)},
                {"url": value(row, 28), "type": value(row, 29)},
            ] if item["url"]],
            "surveyedAt": value(row, 30),
            "confidenceLevel": value(row, 31),
            "needsReviewFlag": value(row, 32),
            "coordinate": value(row, 11),
            "disciplines": split_list(value(row, 12), commas=True),
            "domesticSocieties": split_list(value(row, 13)),
            "internationalSocieties": split_list(value(row, 14)),
            "domesticJournals": split_list(value(row, 15)),
            "internationalJournals": split_list(value(row, 16)),
            "fullPath": " > ".join(path_parts),
            "questions": [q for q in [value(row, 33), value(row, 34)] if q],
        })

    society_sheet = workbook[next(name for name in workbook.sheetnames if name.strip() == "2_学会から探す")]
    societies = []
    for row_number, row in enumerate(rows(society_sheet), start=1):
        name = value(row, 13) or value(row, 1)
        if not name:
            continue
        url_type = value(row, 11)
        societies.append({
            "id": number_id(row[0], "society", row_number),
            "name": name,
            "kind": value(row, 2),
            "kingdom": value(row, 3),
            "division": value(row, 4),
            "className": value(row, 5),
            "orderName": value(row, 6),
            "family": value(row, 7),
            "disciplines": split_list(value(row, 8), commas=True),
            "relatedFields": split_list(value(row, 9), commas=True),
            "url": value(row, 10),
            "urlType": url_type,
            "connectionStatus": status(url_type),
            "description": value(row, 14),
            "questions": [q for q in [value(row, 15), value(row, 16)] if q],
            "sourceUrl": value(row, 17),
            "memberCountEstimate": value(row, 18),
            "memberCountNote": value(row, 19),
            "memberCountAsOf": value(row, 20),
            "activityLevel": value(row, 21),
            "fieldPosition": value(row, 22),
            "accessibility": value(row, 23),
            "evidenceNote": value(row, 24),
            "verificationStatus": value(row, 25),
        })

    journal_sheet = workbook["3_ジャーナルから探す"]
    journals = []
    for row_number, row in enumerate(rows(journal_sheet), start=1):
        name = value(row, 13) or value(row, 1)
        if not name:
            continue
        url_type = value(row, 11)
        journals.append({
            "id": number_id(row[0], "journal", row_number),
            "name": name,
            "kind": value(row, 2),
            "kingdom": value(row, 3),
            "division": value(row, 4),
            "className": value(row, 5),
            "orderName": value(row, 6),
            "family": value(row, 7),
            "disciplines": split_list(value(row, 8), commas=True),
            "relatedFields": split_list(value(row, 9), commas=True),
            "url": value(row, 10),
            "urlType": url_type,
            "connectionStatus": status(url_type),
            "description": value(row, 14),
            "questions": [q for q in [value(row, 15), value(row, 16)] if q],
            "sourceUrl": value(row, 17),
            "publisher": value(row, 18),
            "foundedYear": value(row, 19),
            "frequency": value(row, 20),
            "activityLevel": value(row, 21),
            "peerReview": value(row, 22),
            "articleTypes": value(row, 23),
            "languages": value(row, 24),
            "openAccess": value(row, 25),
            "beginnerReadability": value(row, 26),
            "publicationPosition": value(row, 27),
            "submissionAccessibility": value(row, 28),
            "indexing": value(row, 29),
            "authorGuidelinesUrl": value(row, 30),
            "evidenceNote": value(row, 31),
            "verificationStatus": value(row, 32),
        })

    legend_targets = {
        "学会から探す": {"会員数目安", "会員数_目安", "会員数補足", "会員数情報時点", "活発さ", "分野内での位置づけ", "参加しやすさ"},
        "ジャーナルから探す": {"発行主体", "刊行・更新の活発さ", "査読有無", "論文種別", "言語", "オープンアクセス", "初学者向けの読みやすさ", "掲載媒体としての位置づけ", "投稿しやすさ"},
    }
    legends = []
    for row in rows(workbook["追加情報凡例"]):
        category, item = value(row, 0), value(row, 1)
        if item not in legend_targets.get(category, set()):
            continue
        legends.append({
            "category": "society" if category == "学会から探す" else "journal",
            "item": item.replace("会員数_目安", "会員数目安"),
            "definition": value(row, 2),
            "criteria": value(row, 3),
        })

    edges = []
    for field in fields:
        for key, resource_type in [
            ("domesticSocieties", "society"),
            ("internationalSocieties", "society"),
            ("domesticJournals", "journal"),
            ("internationalJournals", "journal"),
        ]:
            for resource_name in field[key]:
                edges.append({
                    "from": field["id"],
                    "toName": resource_name,
                    "toType": resource_type,
                    "relation": f"field-{resource_type}",
                    "status": "editorial",
                    "source": source.name,
                })

    write_json(output_dir / "fields.json", fields)
    write_json(output_dir / "societies.json", societies)
    write_json(output_dir / "journals.json", journals)
    write_json(output_dir / "research-graph.json", {"edges": edges})
    write_json(output_dir / "resource-legends.json", legends)
    print(json.dumps({"fields": len(fields), "societies": len(societies), "journals": len(journals), "edges": len(edges), "legends": len(legends)}, ensure_ascii=False))


if __name__ == "__main__":
    main()
