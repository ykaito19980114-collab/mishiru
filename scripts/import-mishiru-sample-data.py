#!/usr/bin/env python3
"""Normalize MISHIRU sample Excel files into isolated JSON masters.

This importer is intentionally separate from the current production-ish
`data/labs.json` and `data/normalized/*.json` pipeline. It writes to
`data/mishiru-sample-normalized/` by default so Phase 2 can verify structural
compatibility without replacing the existing master data.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
from collections import Counter, defaultdict
from datetime import date
from pathlib import Path
from typing import Any
from urllib.parse import urlparse

import openpyxl

TAG_STATUS = "pending_STSMP_protocol"
LAB_SHEET = "研究室リスト_DB"
FIELD_SHEET = "1_研究領域から探す"
SOCIETY_SHEET = "2_学会から探す "
JOURNAL_SHEET = "3_ジャーナルから探す"

REQUIRED_COLUMNS = {
    "labs": [
        "No", "大学名", "学部・研究科・専攻", "研究室名", "教授名・職位", "研究分野・キーワード",
        "研究室URL", "研究内容_引用説明", "扱う問い_1", "扱う問い_2",
    ],
    "fields": [
        "日本語名称", "初心者向け説明", "4_この分野が目指すこと（研究目的）",
        "扱っている主な問い_1", "扱っている主な問い_2",
    ],
    "societies": [
        "学会名", "学会説明_引用説明", "扱う問い_1", "扱う問い_2", "関連研究領域(例)",
        "会員数_目安", "活発さ", "分野内での位置づけ", "参加しやすさ", "URL", "評価確認状態",
    ],
    "journals": [
        "ジャーナル名", "ジャーナル説明_引用説明", "扱う問い_1", "扱う問い_2", "関連研究領域(例)",
        "発行主体", "査読有無", "論文種別", "発行頻度", "オープンアクセス",
        "初学者向けの読みやすさ", "URL", "評価確認状態",
    ],
}

OPTIONAL_EXPECTED_COLUMNS = {
    "labs": ["引用元URL", "取得元", "確認状態", "更新日", "原文", "備考"],
    "fields": ["研究領域名", "扱う問い_1", "扱う問い_2", "上位領域", "下位領域", "代表的な研究方法", "関連研究室", "関連学会", "関連ジャーナル"],
    "societies": ["英語名", "初心者向け説明", "主な研究領域", "主なテーマ", "大会・研究会", "公式URL", "確認状態"],
    "journals": ["英語名", "初心者向け説明", "主な研究領域", "査読", "読みやすさ", "公式URL", "確認状態"],
}

PRESERVED_COLUMNS = {
    "labs": [
        "No", "大学名", "学部・研究科・専攻", "研究室名", "教授名・職位", "研究分野・キーワード",
        "研究室URL", "研究内容_引用説明", "扱う問い_1", "扱う問い_2", "研究内容_取得元URL",
        "問い生成根拠メモ", "問い生成確認状態", "URL取得ステータス", "言い過ぎ確認",
        "引用元URL", "取得元", "確認状態", "更新日", "原文", "備考",
    ],
    "fields": [
        "No.", "界", "門", "綱", "目", "科・属", "種", "日本語名称", "英語名称", "階層",
        "定義・学問的立ち位置", "座標(対象×問い)", "対応ディシプリン", "主な国内学会",
        "主な国際学会", "主な国内誌", "主な国際誌", "旧：扱う問い_1", "旧：扱う問い_2",
        "初心者向け説明", "1_標準的な定義", "2_研究対象", "3_代表的な研究方法",
        "4_この分野が目指すこと（研究目的）", "5_代表テーマ", "6_隣接分野との差異",
        "7_根拠資料1", "資料1種別", "7_根拠資料2", "資料2種別", "8_調査日",
        "9_確信度", "9_要確認フラグ", "扱っている主な問い_1", "扱っている主な問い_2",
    ],
    "societies": [
        "No.", "学会名", "種別", "界", "門", "綱", "目", "科・属・種", "対応ディシプリン",
        "関連研究領域(例)", "URL", "URL種別", "No", "学会説明_引用説明", "扱う問い_1",
        "扱う問い_2", "説明取得元URL", "会員数_目安", "会員数補足", "会員数情報時点",
        "活発さ", "分野内での位置づけ", "参加しやすさ", "評価根拠メモ", "評価確認状態",
        "英語名", "初心者向け説明", "主な研究領域", "主なテーマ", "大会・研究会", "公式URL",
        "確認状態",
    ],
    "journals": [
        "No.", "ジャーナル名", "種別", "界", "門", "綱", "目", "科・属・種",
        "対応ディシプリン", "関連研究領域(例)", "URL", "URL種別", "No",
        "ジャーナル説明_引用説明", "扱う問い_1", "扱う問い_2", "説明取得元URL",
        "発行主体", "創刊年", "発行頻度", "刊行・更新の活発さ", "査読有無", "論文種別",
        "言語", "オープンアクセス", "初学者向けの読みやすさ", "掲載媒体としての位置づけ",
        "投稿しやすさ", "収録DB・流通", "投稿規定URL", "評価根拠メモ", "評価確認状態",
        "英語名", "初心者向け説明", "主な研究領域", "査読", "読みやすさ", "公式URL",
        "確認状態",
    ],
}


def clean_header(value: Any) -> str:
    return text(value).replace("\ufeff", "")


def text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def split_list(value: Any, *, commas: bool = True) -> list[str]:
    raw = text(value)
    if not raw:
        return []
    pattern = r"\s*(?:／|/|\n|；|;|、|，|,)\s*" if commas else r"\s*(?:／|/|\n|；|;)\s*"
    return list(dict.fromkeys(part.strip() for part in re.split(pattern, raw) if part.strip()))


def stable_id(prefix: str, *parts: str) -> str:
    raw = "|".join(text(part) for part in parts)
    digest = hashlib.sha1(raw.encode("utf-8")).hexdigest()[:12]
    return f"sample-{prefix}-{digest}"


def valid_url(value: str) -> bool:
    if not value:
        return True
    try:
        parsed = urlparse(value)
        return parsed.scheme in {"http", "https"} and bool(parsed.netloc)
    except ValueError:
        return False


def normalize_url(value: Any) -> str:
    raw = text(value)
    return raw if valid_url(raw) else ""


def connection_status(url_type: str) -> str:
    return "official" if "公式" in url_type else "unverified"


def headers_with_duplicates(header_row: tuple[Any, ...]) -> list[str]:
    counts: Counter[str] = Counter()
    headers: list[str] = []
    for raw in header_row:
        base = clean_header(raw)
        if not base:
            headers.append("")
            continue
        counts[base] += 1
        headers.append(base if counts[base] == 1 else f"{base}__{counts[base]}")
    return headers


def read_records(workbook: openpyxl.Workbook, sheet_name: str) -> tuple[list[str], list[dict[str, str]]]:
    if sheet_name not in workbook.sheetnames:
        raise KeyError(f"Sheet not found: {sheet_name}")
    sheet = workbook[sheet_name]
    rows = sheet.iter_rows(values_only=True)
    header_row = next(rows)
    headers = headers_with_duplicates(header_row)
    records: list[dict[str, str]] = []
    for row_number, row in enumerate(rows, start=2):
        if not any(text(value) for value in row):
            continue
        record = {headers[i]: text(row[i]) if i < len(row) else "" for i in range(len(headers)) if headers[i]}
        record["_rowNumber"] = str(row_number)
        records.append(record)
    return headers, records


def first(row: dict[str, str], *names: str) -> str:
    for name in names:
        if row.get(name):
            return row[name]
    return ""


def graduate_major(department: str) -> tuple[str, str]:
    match = re.search(r"[ 　]", department)
    if not match:
        return department, ""
    return department[: match.start()].strip(), department[match.end() :].strip()


def parse_member(raw: str) -> dict[str, str]:
    value = text(raw)
    if not value:
        return {"name": "", "title": ""}
    titles = ["特任教授", "特命教授", "名誉教授", "招へい教授", "客員教授", "特任准教授", "准教授", "教授", "特任講師", "講師", "特任助教", "助教"]
    title = next((item for item in titles if item in value), "")
    name = value
    for item in titles:
        name = name.replace(item, "")
    name = re.sub(r"[（(].*?[）)]", "", name).strip(" 　・／/")
    return {"name": name, "title": title}


def raw_subset(row: dict[str, str], keys: list[str]) -> dict[str, str]:
    return {key: row.get(key, "") for key in keys if key in row and row.get(key, "")}


def build_mapping_report(headers_by_target: dict[str, list[str]]) -> dict[str, Any]:
    result: dict[str, Any] = {}
    for target, headers in headers_by_target.items():
        cleaned = {h.replace("__2", "").replace("__3", "") for h in headers if h}
        expected = REQUIRED_COLUMNS[target]
        optional = OPTIONAL_EXPECTED_COLUMNS.get(target, [])
        result[target] = {
            "mappedRequiredColumns": [column for column in expected if column in cleaned or (target == "labs" and column == "No" and "\ufeffNo" in headers)],
            "missingRequiredColumns": [column for column in expected if column not in cleaned and not (target == "labs" and column == "No" and "\ufeffNo" in headers)],
            "missingOptionalColumns": [column for column in optional if column not in cleaned],
            "unmappedSourceColumns": [
                h for h in headers
                if h and h.replace("__2", "").replace("__3", "") not in set(PRESERVED_COLUMNS[target])
            ],
            "preservedSourceColumns": [h for h in headers if h and h.replace("__2", "").replace("__3", "") in set(PRESERVED_COLUMNS[target])],
            "duplicateHeaders": [h for h in headers if "__" in h],
        }
    return result


def warn_url(warnings: list[dict[str, str]], entity: str, row: dict[str, str], key: str, value: str) -> None:
    if value and not valid_url(value):
        warnings.append({
            "entity": entity,
            "rowNumber": row.get("_rowNumber", ""),
            "type": "invalid_url",
            "field": key,
            "value": value,
        })


def normalize_labs(records: list[dict[str, str]], warnings: list[dict[str, str]]) -> list[dict[str, Any]]:
    labs: list[dict[str, Any]] = []
    today = date.today().isoformat()
    for row in records:
        source_no = first(row, "No", "\ufeffNo")
        university = first(row, "大学名")
        lab_name = first(row, "研究室名__2", "研究室名") or "研究室"
        department = first(row, "学部・研究科・専攻")
        if not source_no and not university and not lab_name:
            continue
        grad, major = graduate_major(department)
        pi = parse_member(first(row, "教授名・職位"))
        members = [pi] if pi["name"] or pi["title"] else []
        source_keywords = split_list(first(row, "研究分野・キーワード"))
        official_raw = first(row, "研究室URL")
        source_raw = first(row, "研究内容_取得元URL", "引用元URL")
        warn_url(warnings, "lab", row, "研究室URL", official_raw)
        warn_url(warnings, "lab", row, "研究内容_取得元URL", source_raw)
        official_url = normalize_url(official_raw)
        source_url = normalize_url(source_raw)
        sources = []
        if official_url:
            sources.append({"label": "研究室公式サイト", "url": official_url})
        if source_url and source_url != official_url:
            sources.append({"label": "研究内容取得元URL", "url": source_url})
        questions = [q for q in [first(row, "扱う問い_1"), first(row, "扱う問い_2")] if q]
        description = first(row, "研究内容_引用説明")
        labs.append({
            "id": stable_id("lab", source_no, university, lab_name),
            "sourceNo": source_no,
            "sourceSheet": LAB_SHEET,
            "name": lab_name,
            "university": {"name": university, "prefecture": "", "region": ""},
            "university_type": None,
            "department": department,
            "graduate_school": grad,
            "major": major,
            "members": members,
            "pi": pi if pi["name"] or pi["title"] else {"name": "", "title": ""},
            "member_count": len(members) or 1,
            "keywords": source_keywords,
            "sourceKeywords": source_keywords,
            "tags": [],
            "tag_generation_status": TAG_STATUS,
            "area_tags": [],
            "field_major": "other",
            "official_url": official_url or None,
            "has_url": bool(official_url),
            "sources": sources,
            "researchQuestions": questions,
            "questions": questions,
            "sections": {
                "research_summary": description or None,
                "student_themes": None,
                "methods": None,
                "key_papers": None,
                "daily_life": None,
                "mentoring": None,
                "careers": None,
                "fit": None,
                "collaboration": None,
            },
            "status": "published",
            "verified": False,
            "confidence": "public_info",
            "last_updated": first(row, "更新日") or today,
            "quality": {
                "generationBasisNote": first(row, "問い生成根拠メモ"),
                "verificationStatus": first(row, "問い生成確認状態", "確認状態"),
                "urlStatus": first(row, "URL取得ステータス", "URL取得ステータス__2"),
                "overclaimCheck": first(row, "言い過ぎ確認"),
            },
            "rawSource": raw_subset(row, [
                "No", "大学名", "学部・研究科・専攻", "研究室名", "教授名・職位", "研究分野・キーワード",
                "研究室URL", "研究内容_引用説明", "扱う問い_1", "扱う問い_2", "研究内容_取得元URL",
                "問い生成根拠メモ", "問い生成確認状態", "URL取得ステータス", "言い過ぎ確認",
            ]),
        })
    return labs


def normalize_fields(records: list[dict[str, str]]) -> list[dict[str, Any]]:
    fields: list[dict[str, Any]] = []
    for row in records:
        source_no = first(row, "No.")
        name = first(row, "日本語名称", "研究領域名")
        if not name:
            continue
        hierarchy = [first(row, key) for key in ["界", "門", "綱", "目", "科・属"]]
        path_parts = list(dict.fromkeys([part for part in [*hierarchy, name] if part]))
        questions = [q for q in [
            first(row, "扱っている主な問い_1", "扱う問い_1", "旧：扱う問い_1"),
            first(row, "扱っている主な問い_2", "扱う問い_2", "旧：扱う問い_2"),
        ] if q]
        research_objects = split_list(first(row, "2_研究対象"))
        themes = split_list(first(row, "5_代表テーマ"))
        fields.append({
            "id": stable_id("field", source_no, name),
            "sourceNo": source_no,
            "sourceSheet": FIELD_SHEET,
            "nameJa": name,
            "nameEn": first(row, "英語名称"),
            "kingdom": hierarchy[0],
            "division": hierarchy[1],
            "className": hierarchy[2],
            "orderName": hierarchy[3],
            "family": hierarchy[4],
            "level": first(row, "階層"),
            "definition": first(row, "1_標準的な定義", "定義・学問的立ち位置"),
            "beginnerDescription": first(row, "初心者向け説明"),
            "researchObjects": research_objects,
            "methods": split_list(first(row, "3_代表的な研究方法")),
            "researchPurpose": first(row, "4_この分野が目指すこと（研究目的）"),
            "representativeThemes": themes,
            "adjacentDifference": first(row, "6_隣接分野との差異"),
            "evidenceSources": [
                {"url": first(row, "7_根拠資料1"), "type": first(row, "資料1種別")},
                {"url": first(row, "7_根拠資料2"), "type": first(row, "資料2種別")},
            ],
            "surveyedAt": first(row, "8_調査日"),
            "confidenceLevel": first(row, "9_確信度"),
            "needsReviewFlag": first(row, "9_要確認フラグ"),
            "coordinate": first(row, "座標(対象×問い)"),
            "disciplines": split_list(first(row, "対応ディシプリン")),
            "domesticSocieties": split_list(first(row, "主な国内学会"), commas=False),
            "internationalSocieties": split_list(first(row, "主な国際学会"), commas=False),
            "domesticJournals": split_list(first(row, "主な国内誌"), commas=False),
            "internationalJournals": split_list(first(row, "主な国際誌"), commas=False),
            "fullPath": " > ".join(path_parts),
            "questions": questions,
            "sourceKeywords": list(dict.fromkeys([*research_objects, *themes, *split_list(first(row, "対応ディシプリン"))])),
            "tags": [],
            "tag_generation_status": TAG_STATUS,
        })
    return fields


def normalize_societies(records: list[dict[str, str]], warnings: list[dict[str, str]]) -> list[dict[str, Any]]:
    societies: list[dict[str, Any]] = []
    for row in records:
        source_no = first(row, "No.")
        name = first(row, "学会名__2", "学会名")
        if not name:
            continue
        url_raw = first(row, "URL", "公式URL")
        source_raw = first(row, "説明取得元URL")
        warn_url(warnings, "society", row, "URL", url_raw)
        warn_url(warnings, "society", row, "説明取得元URL", source_raw)
        related_fields = split_list(first(row, "関連研究領域(例)", "主な研究領域"))
        disciplines = split_list(first(row, "対応ディシプリン"))
        societies.append({
            "id": stable_id("society", source_no, name),
            "sourceNo": source_no,
            "sourceSheet": SOCIETY_SHEET,
            "name": name,
            "nameEn": first(row, "英語名"),
            "kind": first(row, "種別"),
            "kingdom": first(row, "界"),
            "division": first(row, "門"),
            "className": first(row, "綱"),
            "orderName": first(row, "目"),
            "family": first(row, "科・属・種"),
            "disciplines": disciplines,
            "relatedFields": related_fields,
            "url": normalize_url(url_raw),
            "urlType": first(row, "URL種別"),
            "connectionStatus": connection_status(first(row, "URL種別")),
            "description": first(row, "初心者向け説明", "学会説明_引用説明"),
            "beginnerDescription": first(row, "初心者向け説明", "学会説明_引用説明"),
            "questions": [q for q in [first(row, "扱う問い_1"), first(row, "扱う問い_2")] if q],
            "sourceUrl": normalize_url(source_raw),
            "memberCountEstimate": first(row, "規模", "会員数_目安"),
            "memberCountNote": first(row, "会員数補足"),
            "memberCountAsOf": first(row, "会員数情報時点"),
            "activityLevel": first(row, "活発さ"),
            "fieldPosition": first(row, "位置づけ", "分野内での位置づけ"),
            "accessibility": first(row, "初学者・社会人の参加しやすさ", "参加しやすさ"),
            "meetingInfo": first(row, "大会・研究会"),
            "evidenceNote": first(row, "評価根拠メモ"),
            "verificationStatus": first(row, "確認状態", "評価確認状態"),
            "sourceKeywords": list(dict.fromkeys([*related_fields, *disciplines])),
            "tags": [],
            "tag_generation_status": TAG_STATUS,
        })
    return societies


def normalize_journals(records: list[dict[str, str]], warnings: list[dict[str, str]]) -> list[dict[str, Any]]:
    journals: list[dict[str, Any]] = []
    for row in records:
        source_no = first(row, "No.")
        name = first(row, "ジャーナル名__2", "ジャーナル名")
        if not name:
            continue
        url_raw = first(row, "URL", "公式URL")
        source_raw = first(row, "説明取得元URL")
        guideline_raw = first(row, "投稿規定URL")
        warn_url(warnings, "journal", row, "URL", url_raw)
        warn_url(warnings, "journal", row, "説明取得元URL", source_raw)
        warn_url(warnings, "journal", row, "投稿規定URL", guideline_raw)
        related_fields = split_list(first(row, "関連研究領域(例)", "主な研究領域"))
        disciplines = split_list(first(row, "対応ディシプリン"))
        journals.append({
            "id": stable_id("journal", source_no, name),
            "sourceNo": source_no,
            "sourceSheet": JOURNAL_SHEET,
            "name": name,
            "nameEn": first(row, "英語名"),
            "kind": first(row, "種別"),
            "kingdom": first(row, "界"),
            "division": first(row, "門"),
            "className": first(row, "綱"),
            "orderName": first(row, "目"),
            "family": first(row, "科・属・種"),
            "disciplines": disciplines,
            "relatedFields": related_fields,
            "url": normalize_url(url_raw),
            "urlType": first(row, "URL種別"),
            "connectionStatus": connection_status(first(row, "URL種別")),
            "description": first(row, "初心者向け説明", "ジャーナル説明_引用説明"),
            "beginnerDescription": first(row, "初心者向け説明", "ジャーナル説明_引用説明"),
            "questions": [q for q in [first(row, "扱う問い_1"), first(row, "扱う問い_2")] if q],
            "sourceUrl": normalize_url(source_raw),
            "publisher": first(row, "発行主体"),
            "foundedYear": first(row, "創刊年"),
            "frequency": first(row, "発行頻度"),
            "activityLevel": first(row, "刊行・更新の活発さ"),
            "peerReview": first(row, "査読", "査読有無"),
            "articleTypes": first(row, "論文種別"),
            "languages": first(row, "言語"),
            "openAccess": first(row, "オープンアクセス"),
            "beginnerReadability": first(row, "読みやすさ", "初学者向けの読みやすさ"),
            "publicationPosition": first(row, "掲載媒体としての位置づけ"),
            "submissionAccessibility": first(row, "投稿しやすさ"),
            "indexing": first(row, "収録DB・流通"),
            "authorGuidelinesUrl": normalize_url(guideline_raw),
            "evidenceNote": first(row, "評価根拠メモ"),
            "verificationStatus": first(row, "確認状態", "評価確認状態"),
            "sourceKeywords": list(dict.fromkeys([*related_fields, *disciplines])),
            "tags": [],
            "tag_generation_status": TAG_STATUS,
        })
    return journals


def duplicate_report(items: list[dict[str, Any]], identity_keys: list[str]) -> dict[str, Any]:
    id_counts = Counter(item["id"] for item in items)
    identity_counts = Counter("|".join(text(item.get(key, "")) for key in identity_keys) for item in items)
    return {
        "idDuplicates": sum(count - 1 for count in id_counts.values() if count > 1),
        "identityDuplicates": sum(count - 1 for count in identity_counts.values() if count > 1 and any(identity_counts)),
        "duplicateIdentities": [key for key, count in identity_counts.items() if count > 1 and key.strip("|")][:20],
    }


def missing_report(items: list[dict[str, Any]], keys: list[str]) -> dict[str, int]:
    report: dict[str, int] = {}
    for key in keys:
        report[key] = sum(1 for item in items if not item.get(key))
    return report


def write_json(path: Path, payload: Any) -> None:
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--labs", type=Path, required=True)
    parser.add_argument("--resources", type=Path, required=True)
    parser.add_argument("--out", type=Path, default=Path("data/mishiru-sample-normalized"))
    args = parser.parse_args()

    args.out.mkdir(parents=True, exist_ok=True)
    lab_book = openpyxl.load_workbook(args.labs, read_only=True, data_only=True)
    resource_book = openpyxl.load_workbook(args.resources, read_only=True, data_only=True)

    lab_headers, lab_records = read_records(lab_book, LAB_SHEET)
    field_headers, field_records = read_records(resource_book, FIELD_SHEET)
    society_headers, society_records = read_records(resource_book, SOCIETY_SHEET)
    journal_headers, journal_records = read_records(resource_book, JOURNAL_SHEET)

    warnings: list[dict[str, str]] = []
    labs = normalize_labs(lab_records, warnings)
    fields = normalize_fields(field_records)
    societies = normalize_societies(society_records, warnings)
    journals = normalize_journals(journal_records, warnings)

    outputs = {
        "labs.json": labs,
        "fields.json": fields,
        "societies.json": societies,
        "journals.json": journals,
    }
    for filename, payload in outputs.items():
        write_json(args.out / filename, payload)

    headers_by_target = {
        "labs": lab_headers,
        "fields": field_headers,
        "societies": society_headers,
        "journals": journal_headers,
    }
    mapping = build_mapping_report(headers_by_target)
    report = {
        "sources": {
            "labs": str(args.labs),
            "resources": str(args.resources),
        },
        "sheets": {
            "labs": LAB_SHEET,
            "fields": FIELD_SHEET,
            "societies": SOCIETY_SHEET,
            "journals": JOURNAL_SHEET,
        },
        "inputRows": {
            "labs": len(lab_records),
            "fields": len(field_records),
            "societies": len(society_records),
            "journals": len(journal_records),
        },
        "normalizedRows": {
            "labs": len(labs),
            "fields": len(fields),
            "societies": len(societies),
            "journals": len(journals),
        },
        "duplicates": {
            "labs": duplicate_report(labs, ["name", "department"]),
            "fields": duplicate_report(fields, ["nameJa"]),
            "societies": duplicate_report(societies, ["name"]),
            "journals": duplicate_report(journals, ["name"]),
        },
        "missing": {
            "labs": missing_report(labs, ["name", "department", "official_url", "researchQuestions"]),
            "fields": missing_report(fields, ["nameJa", "beginnerDescription", "researchPurpose", "questions"]),
            "societies": missing_report(societies, ["name", "description", "questions", "url"]),
            "journals": missing_report(journals, ["name", "description", "questions", "url"]),
        },
        "warnings": {
            "total": len(warnings),
            "invalidUrls": sum(1 for warning in warnings if warning["type"] == "invalid_url"),
        },
        "columnMapping": mapping,
        "tagPolicy": {
            "tags": [],
            "tag_generation_status": TAG_STATUS,
            "note": "sourceKeywords are preserved separately from future STSMP tags.",
        },
    }
    write_json(args.out / "import-report.json", report)
    write_json(args.out / "import-warnings.json", warnings)
    write_json(args.out / "column-mapping-report.json", mapping)
    print(json.dumps(report, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
