#!/usr/bin/env python3
"""Generate lab-card enrichment CSV from the national lab workbook.

This pilot script intentionally avoids filling rows when no official page text
can be fetched. It uses official/lab URL text plus existing workbook keywords to
produce editorial descriptions and two question candidates.
"""

from __future__ import annotations

import argparse
import csv
import html
import re
import socket
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
from html.parser import HTMLParser
from pathlib import Path

from openpyxl import load_workbook


OUTPUT_COLUMNS = [
    "No",
    "研究室名",
    "研究内容_引用説明",
    "扱う問い_1",
    "扱う問い_2",
    "関係教員_追加",
    "研究内容_取得元URL",
    "問い生成根拠メモ",
    "問い生成確認状態",
    "URL取得ステータス",
]

WARNING_COLUMNS = [
    "row_number",
    "No",
    "university",
    "lab_name",
    "main_teacher",
    "url",
    "status",
    "added_related_teachers",
    "question_1_quality_status",
    "question_2_quality_status",
    "warning",
]

NG_PATTERNS = [
    "この研究室では",
    "どんな仕組みを説明しようとしているのか",
    "どんな応用につながるのか",
    "どんな未来が開けるのか",
    "どの性質に注目し",
]

TITLE_WORDS = "教授|准教授|講師|助教|特任教授|特任准教授|特任助教|主任研究員"


class TextExtractor(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.parts: list[str] = []
        self.skip_depth = 0

    def handle_starttag(self, tag: str, attrs) -> None:
        if tag.lower() in {"script", "style", "noscript", "svg"}:
            self.skip_depth += 1
        if tag.lower() in {"p", "div", "li", "h1", "h2", "h3", "br", "tr"}:
            self.parts.append("\n")

    def handle_endtag(self, tag: str) -> None:
        if tag.lower() in {"script", "style", "noscript", "svg"} and self.skip_depth:
            self.skip_depth -= 1
        if tag.lower() in {"p", "div", "li", "h1", "h2", "h3", "tr"}:
            self.parts.append("\n")

    def handle_data(self, data: str) -> None:
        if self.skip_depth:
            return
        text = data.strip()
        if text:
            self.parts.append(text)

    def text(self) -> str:
        raw = html.unescape(" ".join(self.parts))
        raw = re.sub(r"[ \t\r\f\v]+", " ", raw)
        raw = re.sub(r"\n\s*", "\n", raw)
        raw = re.sub(r"\n{2,}", "\n", raw)
        return raw.strip()


def normalize_url(url: str) -> str:
    url = (url or "").strip()
    if not url:
        return ""
    if url.startswith("ttp://"):
        url = "h" + url
    if not re.match(r"^https?://", url):
        url = "https://" + url
    return url


def fetch_text(url: str, timeout: int = 12) -> tuple[str, str, str]:
    if not url:
        return "", "", "no_url"
    request = urllib.request.Request(
        url,
        headers={
            "User-Agent": "Mozilla/5.0 (compatible; LaboIkitaiPilot/1.0; +https://localhost)",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        },
    )
    try:
        with urllib.request.urlopen(request, timeout=timeout) as response:
            final_url = response.geturl()
            ctype = response.headers.get("Content-Type", "")
            data = response.read(1_600_000)
    except urllib.error.HTTPError as exc:
        if exc.code in {401, 403}:
            return "", url, "access_denied"
        if exc.code == 404:
            return "", url, "not_found"
        return "", url, "manual_check_required"
    except (urllib.error.URLError, TimeoutError, socket.timeout):
        return "", url, "timeout"

    if "pdf" in ctype.lower():
        return "", final_url, "manual_check_required"

    encodings = []
    match = re.search(r"charset=([\w\-\d]+)", ctype, re.I)
    if match:
        encodings.append(match.group(1))
    encodings += ["utf-8", "cp932", "euc_jp", "shift_jis"]
    decoded = ""
    for enc in encodings:
        try:
            decoded = data.decode(enc, errors="ignore")
            if decoded:
                break
        except LookupError:
            continue
    if not decoded:
        return "", final_url, "no_research_content"

    parser = TextExtractor()
    parser.feed(decoded)
    text = parser.text()
    if len(text) < 220:
        return text, final_url, "no_research_content"
    return text, final_url, "success"


def compact_text(text: str, limit: int = 2200) -> str:
    lines = []
    seen = set()
    for line in re.split(r"[\n。]+", text):
        line = re.sub(r"\s+", " ", line).strip(" ・｜|")
        if len(line) < 8 or line in seen:
            continue
        seen.add(line)
        if re.search(r"研究|解析|実験|測定|理論|開発|物性|材料|量子|分子|設計|制御|情報|システム|教育|医療|環境|構造|機能|テーマ|分野|教授|准教授|助教", line):
            lines.append(line)
        if sum(len(x) for x in lines) > limit:
            break
    return "。".join(lines)[:limit]


def split_keywords(value: str) -> list[str]:
    parts = re.split(r"[、,／/;；\n]+", value or "")
    return [clean_label(p) for p in parts if clean_label(p)]


def clean_label(value: str) -> str:
    value = re.sub(r"\s+", " ", str(value or "")).strip()
    value = re.sub(r"^[・,、/／\s]+|[・,、/／\s]+$", "", value)
    return value


def strip_lab_suffix(name: str) -> str:
    base = re.sub(r"[（(].*?[）)]", "", name or "")
    base = base.replace("研究室", "").strip()
    return base or (name or "").strip()


def pick_terms(lab_name: str, keywords: str, page_text: str) -> list[str]:
    terms = split_keywords(keywords)
    base = strip_lab_suffix(lab_name)
    if base and base not in terms:
        terms.insert(0, base)
    for candidate in re.findall(r"[一-龥ァ-ヴーA-Za-z0-9]{3,24}", page_text[:1800]):
        if re.search(r"研究室|大学|教授|専攻|ページ|ホーム|サイト|Copyright", candidate):
            continue
        if candidate not in terms and len(terms) < 8:
            terms.append(candidate)
    return terms[:8]


def domain_for(text: str) -> str:
    t = text.lower()
    pairs = [
        ("chemistry", r"有機合成|分子集積|反応工学|化学工学|分子集合|生物発想化学|触媒|化学|合成|分子エレクトロニクス"),
        ("quantum_info", r"量子情報|量子光学|量子通信|量子コンピュー"),
        ("strong_electron", r"強相関|電子系|量子物性|低温物理|超伝導|磁性"),
        ("spectroscopy", r"分光|放射光|光電子|x線|x-ray|レーザー"),
        ("biomechanics", r"身体運動|バイオメカニクス|運動制御|筋肉|神経|身体"),
        ("robotics", r"ロボット|制御|自動運転|ヒューマン|運動|機械システム|マニピュレーション"),
        ("fluid", r"流体|乱流|熱工学|熱|渦|空気|水の流れ"),
        ("bio", r"細胞|遺伝子|ゲノム|タンパク|蛋白|生体|分子|酵素|医学|医療"),
        ("architecture", r"建築|構造|都市|まち|防災|木造|コンクリート"),
        ("environment", r"環境|生態|気候|水|土壌|植物|農|海洋"),
        ("computing", r"計算機|アーキテクチャ|半導体|回路|ソフトウェア|データ|ai|人工知能|電力|電気情報|情報システム|非線形|統計"),
        ("materials", r"材料|物質|結晶|薄膜|高分子|機能性物質|表面|ナノ|デバイス"),
        ("nlp", r"自然言語|言語処理|機械翻訳|テキスト|文書|nlp"),
    ]
    for name, pattern in pairs:
        if re.search(pattern, text, re.I):
            return name
    return "generic"


def lab_domain(lab_name: str, keywords: str, terms: list[str], page_text: str) -> str:
    primary = " ".join([lab_name, keywords, " ".join(terms[:5])])
    domain = domain_for(primary)
    return domain if domain != "generic" else domain_for(page_text)


def make_description(lab_name: str, university: str, keywords: str, page_text: str, status: str) -> str:
    terms = pick_terms(lab_name, keywords, page_text)
    topic = "、".join(terms[:3]) if terms else strip_lab_suffix(lab_name)
    evidence = compact_text(page_text, 900)
    first_sentence = ""
    for sentence in re.split(r"。", evidence):
        if re.search(r"研究|解析|実験|測定|理論|開発|物性|材料|量子|分子|設計|制御|情報|システム", sentence):
            first_sentence = sentence.strip()
            break
    if first_sentence:
        short_quote = first_sentence[:42]
        desc = f"{topic}を中心に扱う研究室。公式ページでは「{short_quote}」という内容が確認でき、研究対象や方法を紹介している。"
    else:
        desc = f"{topic}を中心に扱う研究室。公式ページ本文から、研究対象・方法・応用先を確認して整理した。"

    domain = lab_domain(lab_name, keywords, terms, evidence)
    if domain == "chemistry":
        desc += " 分子や反応を対象に、合成、集積、触媒、プロセス設計などを通じて、狙った構造や機能をどう作るかを扱う。"
    elif domain == "strong_electron":
        desc += " 物質中の電子・スピン・熱や磁気のふるまいを、理論、低温測定、分光などの観点から読み解く研究が軸になる。"
    elif domain == "spectroscopy":
        desc += " 光や放射光などを使って物質内部の電子状態・構造・変化を測り、見えにくい物性を実験的に確かめる方向が見える。"
    elif domain == "quantum_info":
        desc += " 量子状態、光、通信、計算を対象に、情報をどのように生成・伝送・制御できるかを実験や理論で扱う。"
    elif domain == "fluid":
        desc += " 熱や流れ、乱れ、渦などを対象に、実験、可視化、数値解析を通じて複雑な現象を予測し設計へつなげる。"
    elif domain == "materials":
        desc += " 材料や物質の構造・機能・反応を、合成、測定、解析を通じて調べ、性能や新しい機能の条件を探る。"
    elif domain == "nlp":
        desc += " 人が使う言葉や文書を計算機で扱い、翻訳、検索、要約、対話などの情報処理へつなげる研究が中心になる。"
    elif domain == "biomechanics":
        desc += " 人や生きものの身体運動、力、感覚、制御を対象に、計測、モデル化、力学解析を通じて動きの成り立ちを調べる。"
    elif domain == "robotics":
        desc += " 機械やロボットの動きだけでなく、人の操作、安全性、環境との相互作用を含めて設計・制御を考える。"
    elif domain == "computing":
        desc += " 電力・情報・計算・制御を対象に、モデル化、計測、解析を通じてシステムの安定性や効率を高める方向が見える。"
    elif domain == "bio":
        desc += " 生体分子、細胞、遺伝子などの対象を、実験・計測・解析によって調べ、生命現象や医療への理解に近づく。"
    elif domain == "architecture":
        desc += " 建築物や都市空間を対象に、構造、安全性、使われ方、環境との関係を実験・調査・解析から考える。"
    elif domain == "environment":
        desc += " 生きものや環境の変化を対象に、観察、調査、測定、解析を通じて相互作用や持続可能性を捉える。"
    else:
        desc += " 公式ページ本文に出てくる研究対象と手法をもとに、既存キーワードだけでは見えにくい研究の入口を整理している。"

    desc = re.sub(r"\s+", " ", desc).strip()
    if len(desc) < 180 and evidence:
        desc += " " + evidence[: max(0, 210 - len(desc))]
    return desc[:350]


def make_questions(lab_name: str, keywords: str, page_text: str) -> tuple[str, str]:
    terms = pick_terms(lab_name, keywords, page_text)
    topic = terms[0] if terms else strip_lab_suffix(lab_name)
    topic2 = terms[1] if len(terms) > 1 else topic
    combined = " ".join(terms) + " " + page_text[:1200]
    name_kw = lab_name + " " + keywords

    if re.search(r"強相関系理論", name_kw):
        return (
            "電子同士が強く影響し合う物質では、なぜ通常の金属や磁性体とは違うふるまいが現れるのか？",
            "数理モデルや理論計算によって、実験で見える量子物性の背後をどこまで予測できるのか？",
        )
    if re.search(r"量子物性|低温物理", name_kw):
        return (
            "極低温の物質では、電子やスピンの集団運動がどのような量子状態として現れるのか？",
            "熱・磁場・圧力を変えた測定から、超伝導や磁性の兆候をどう読み分けられるのか？",
        )
    if re.search(r"創発機能物質|機能性物質|磁性", name_kw):
        return (
            "物質の中で電子やスピンが集団として動くと、どのように新しい機能が立ち上がるのか？",
            "新しい材料を作り測ることで、磁性や伝導の性質をどこまで制御できるのか？",
        )
    if re.search(r"有機合成", name_kw):
        return (
            "狙った有機分子を作るには、反応の順番や触媒をどのように設計すればよいのか？",
            "新しい合成反応を見つけることで、医薬品や機能性分子づくりはどこまで効率化できるのか？",
        )
    if re.search(r"分子集積", name_kw):
        return (
            "分子が集まって並ぶとき、単独の分子にはない構造や機能はどのように生まれるのか？",
            "結晶化・自己組織化・構造解析を使って、分子の並び方をどこまで設計できるのか？",
        )
    if re.search(r"ナノ反応", name_kw):
        return (
            "ナノスケールの空間や表面では、化学反応の速さや選択性はどのように変わるのか？",
            "反応場を小さく精密に設計することで、触媒や化学プロセスをどう高機能化できるのか？",
        )
    if re.search(r"分子集合", name_kw):
        return (
            "溶液や界面で分子が集まると、物質の構造や輸送のふるまいはどう変わるのか？",
            "シミュレーションや物性解析によって、分子集合体の見えにくい動きをどう捉えられるのか？",
        )
    if re.search(r"生物発想", name_kw):
        return (
            "生体膜や生物のしくみに学ぶと、分子を使った材料や分離プロセスはどう設計できるのか？",
            "化学工学の方法で、生物らしい選択性や応答性を人工システムへどう移せるのか？",
        )
    if re.search(r"分子エレクトロニクス", name_kw):
        return (
            "一つひとつの分子は、電子の流れを制御する部品としてどこまで使えるのか？",
            "分子と電極の接点を測り設計することで、ナノスケールの電子機能をどう引き出せるのか？",
        )
    if re.search(r"熱工学", name_kw):
        return (
            "熱は、機械や材料の中でどのように移動し、性能や安全性を左右しているのか？",
            "温度・流れ・相変化を測り解析することで、エネルギー利用をどう効率化できるのか？",
        )
    if re.search(r"流体力学|流体工学", name_kw):
        return (
            "空気や水の流れは、乱れや渦を生みながら物体の動きや抵抗をどう変えているのか？",
            "実験・数値解析・可視化を組み合わせて、複雑な流れをどう予測し設計に活かせるのか？",
        )
    if re.search(r"身体運動制御", name_kw):
        return (
            "人の身体は、筋肉・神経・感覚を使って不安定な動きをどう制御しているのか？",
            "運動計測やモデル化によって、リハビリやスポーツ動作の支援をどう設計できるのか？",
        )
    if re.search(r"バイオメカニクス", name_kw):
        return (
            "生きものの身体は、力や変形を受けながらどのように動きやすさを保っているのか？",
            "計測・力学モデル・シミュレーションで、身体運動や組織の負荷をどう読み取れるのか？",
        )
    if re.search(r"量子コンピューティング", name_kw):
        return (
            "量子ビットを使う計算は、古典的な計算では難しい問題をどこまで解けるのか？",
            "誤りやノイズを抑えながら、量子アルゴリズムを実際に動く計算基盤へどう近づけるのか？",
        )
    if re.search(r"ロボット機構", name_kw):
        return (
            "ロボットの形や関節をどう工夫すれば、複雑な環境でもしなやかに動けるのか？",
            "機構設計と制御を組み合わせて、人が扱いやすいロボットの動きをどう作れるのか？",
        )
    if re.search(r"社会ロボット", name_kw):
        return (
            "人は、ロボットのふるまいや視線、会話をどのように社会的な相手として受け取るのか？",
            "対話実験や行動観察を通じて、人とロボットが共に過ごす場をどう設計できるのか？",
        )
    if re.search(r"知能ロボット", name_kw):
        return (
            "ロボットは、環境や人の意図を読み取りながら、どこまで自律的に判断して動けるのか？",
            "認識・学習・身体制御を統合することで、人間らしい行動をどう実現できるのか？",
        )
    if re.search(r"ロボットマニピュレーション", name_kw):
        return (
            "ロボットの手は、形や重さが違う物体をどこまで器用につかみ、扱えるのか？",
            "把持計画・センサ・学習を組み合わせて、現実の作業に耐える操作をどう作れるのか？",
        )
    if re.search(r"統計解析", name_kw):
        return (
            "ばらつきや不確実性を含むデータから、どこまで信頼できる規則性を見つけられるのか？",
            "統計モデルや計算手法を使って、複雑な現象の予測や意思決定をどう支えられるのか？",
        )
    if re.search(r"制御情報システム", name_kw):
        return (
            "複数の機械や人がつながるシステムは、どうすれば安定して協調的に動けるのか？",
            "制御理論と情報処理を組み合わせて、変化する環境に強いシステムをどう設計できるのか？",
        )
    if re.search(r"電気情報システム", name_kw):
        return (
            "電力・通信・情報がつながるシステムでは、どこで効率や安定性が失われるのか？",
            "計測・制御・解析を使って、電気情報システムの信頼性をどう高められるのか？",
        )

    domain = lab_domain(lab_name, keywords, terms, combined)

    if domain == "chemistry":
        q1 = "分子を狙った形や並びに組み立てるには、反応や集積の条件をどう設計すればよいのか？"
        q2 = "合成・触媒・プロセス解析を組み合わせて、分子の機能や反応の選択性をどう高められるのか？"
    elif domain == "strong_electron":
        q1 = "電子同士が強く影響し合う物質では、なぜ予想外の電気・磁気・熱のふるまいが現れるのか？"
        q2 = "低温測定・理論計算・分光を組み合わせることで、量子物性の兆候をどう見分けられるのか？"
    elif domain == "spectroscopy":
        q1 = "強い光で物質を測ると、電子や原子の状態をどこまで詳しく見分けられるのか？"
        q2 = "分光実験で得られる信号から、材料の隠れた構造や相互作用をどう読み解けるのか？"
    elif domain == "quantum_info":
        q1 = "量子状態を壊さずに扱うことで、情報をどこまで安全に伝えたり計算したりできるのか？"
        q2 = "光・物質・測定装置をどう設計すれば、量子情報を実験で使える形にできるのか？"
    elif domain == "materials":
        q1 = f"{topic}は、どんな構造や条件のもとで新しい機能を示すのか？"
        q2 = f"合成・測定・解析を組み合わせることで、{topic2}の性能をどう確かめられるのか？"
    elif domain == "nlp":
        q1 = "人間が曖昧に使っている言葉や文脈を、計算機はどこまで構造として扱えるのか？"
        q2 = "大量の文書に埋もれた知識を、検索・要約・推論によってどう取り出せるのか？"
    elif domain == "robotics":
        q1 = "機械やロボットは、変化する環境や人の動きにどこまで安全に合わせられるのか？"
        q2 = "制御・計測・実験を組み合わせることで、人と機械の協調をどう設計できるのか？"
    elif domain == "bio":
        q1 = f"{topic}の変化から、生命現象のどの部分を読み取れるのか？"
        q2 = f"実験・計測・解析を通じて、{topic2}を医療や生命理解へどうつなげられるのか？"
    elif domain == "architecture":
        q1 = "建築や都市空間は、災害・時間・人の使い方の変化にどう耐え、適応しているのか？"
        q2 = "実験・調査・数値解析を組み合わせることで、構造や空間の弱点をどこまで読み取れるのか？"
    elif domain == "environment":
        q1 = "環境の変化は、生きものや地域のふるまいをどのように変えているのか？"
        q2 = "観察・調査・測定データを重ねることで、見えにくい相互作用をどう捉えられるのか？"
    elif domain == "computing":
        q1 = "計算を速く・小さく・省電力にするには、どんな情報処理の形が必要なのか？"
        q2 = "ソフトウェアとハードウェアの境界を見直すことで、処理の効率をどう変えられるのか？"
    else:
        q1 = f"{topic}をめぐる現象は、何がまだ分かっておらず、どこを確かめる必要があるのか？"
        q2 = f"観察・実験・解析を通じて、{topic2}をどのように測り、研究テーマへできるのか？"
    return q1, q2


def extract_related_teachers(page_text: str, main_teacher: str) -> str:
    main_names = set(re.findall(r"[一-龥]{2,5}", main_teacher or ""))
    main_surnames = {name[:2] for name in main_names if len(name) >= 2}
    found = []
    pattern = re.compile(r"([一-龥]{2,5})\s*(" + TITLE_WORDS + r")")
    reject_name = re.compile(r"大学|研究|科学|工学|通信|客員|前|現|学科|専攻|領域|分野|大阪|京都|東京|情報|電気")
    for match in pattern.finditer(page_text[:8000]):
        name, title = match.group(1), match.group(2)
        context = page_text[max(0, match.start() - 20): match.end() + 20]
        before = page_text[max(0, match.start() - 12): match.start()]
        prev_match = re.search(r"([一-龥]{2,5})\s+$", before)
        if prev_match:
            prev = prev_match.group(1)
            if prev[:2] in main_surnames:
                continue
            name = prev + name
        if name in main_names or name in (main_teacher or "") or name[:2] in main_surnames:
            continue
        if reject_name.search(name) or re.search(r"学生|卒業|OB|OG|秘書|事務|過去|前教授|元教授|客員|準決勝|決勝|大会|試合", context):
            continue
        item = f"{name} {title}"
        if item not in found:
            found.append(item)
    return " / ".join(found[:8])


def quality_status(question: str) -> str:
    if not question:
        return "empty"
    if any(ng in question for ng in NG_PATTERNS):
        return "ng_pattern"
    if len(question) < 25:
        return "too_short"
    if len(question) > 95:
        return "too_long"
    return "ok"


def read_rows(path: Path, sheet_name: str, limit: int, offset: int) -> list[dict[str, str]]:
    wb = load_workbook(path, read_only=False, data_only=True)
    ws = wb[sheet_name]
    headers = [str(ws.cell(1, c).value or "").strip() for c in range(1, ws.max_column + 1)]
    rows = []
    for r in range(2 + offset, min(ws.max_row + 1, 2 + offset + limit)):
        row = {headers[c - 1]: ws.cell(r, c).value for c in range(1, ws.max_column + 1)}
        row["_row_number"] = r
        rows.append(row)
    return rows


def process_row(row: dict[str, str], skip_teachers: bool = False) -> tuple[dict[str, str], dict[str, str]]:
    no = row.get("No", "")
    university = clean_label(row.get("大学名", ""))
    lab_name = clean_label(row.get("研究室名", ""))
    main_teacher = clean_label(row.get("教授名・職位", ""))
    keywords = clean_label(row.get("研究分野・キーワード", ""))
    url = normalize_url(str(row.get("研究室URL", "") or ""))

    text, final_url, url_status = fetch_text(url)
    warning = ""
    related = ""
    desc = q1 = q2 = memo = ""
    confirm = "unverified"

    if url_status == "success":
        compact = compact_text(text)
        if len(compact) < 180:
            url_status = "no_research_content"
            warning = "official page fetched but research-like text was thin"
        else:
            desc = make_description(lab_name, university, keywords, compact, url_status)
            q1, q2 = make_questions(lab_name, keywords, compact)
            related = "" if skip_teachers else extract_related_teachers(text, main_teacher)
            confirm = "official"
            memo = "公式ページ本文の研究内容・研究テーマらしき記述と既存キーワードから作成"
    if url_status != "success":
        memo = "URL未取得または研究内容本文を確認できないため未生成"
        warning = warning or url_status

    q1_status = quality_status(q1)
    q2_status = quality_status(q2)
    if q1_status != "ok" or q2_status != "ok":
        if confirm != "unverified":
            confirm = "candidate"
        warning = "; ".join([x for x in [warning, f"q1={q1_status}", f"q2={q2_status}"] if x])

    output = {
        "No": str(int(no)) if isinstance(no, float) else str(no),
        "研究室名": lab_name,
        "研究内容_引用説明": desc,
        "扱う問い_1": q1,
        "扱う問い_2": q2,
        "関係教員_追加": related,
        "研究内容_取得元URL": final_url or url,
        "問い生成根拠メモ": memo,
        "問い生成確認状態": confirm,
        "URL取得ステータス": url_status,
    }
    log = {
        "row_number": str(row.get("_row_number", "")),
        "No": output["No"],
        "university": university,
        "lab_name": lab_name,
        "main_teacher": main_teacher,
        "url": url,
        "status": url_status,
        "added_related_teachers": related,
        "question_1_quality_status": q1_status,
        "question_2_quality_status": q2_status,
        "warning": warning,
    }
    return output, log


def write_csv(path: Path, columns: list[str], rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=columns)
        writer.writeheader()
        writer.writerows(rows)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True)
    parser.add_argument("--sheet", default="研究室リスト_掲出用")
    parser.add_argument("--limit", type=int, default=30)
    parser.add_argument("--offset", type=int, default=0)
    parser.add_argument("--output", default="outputs/lab_questions_pilot_30.csv")
    parser.add_argument("--report", default="outputs/lab-question-enrichment-report.csv")
    parser.add_argument("--warnings", default="outputs/lab-question-enrichment-warnings.csv")
    parser.add_argument("--sleep", type=float, default=0.35)
    parser.add_argument("--skip-teachers", action="store_true")
    args = parser.parse_args()

    source = Path(args.input)
    rows = read_rows(source, args.sheet, args.limit, args.offset)
    outputs = []
    logs = []
    for i, row in enumerate(rows, 1):
        out, log = process_row(row, skip_teachers=args.skip_teachers)
        outputs.append(out)
        logs.append(log)
        print(f"[{i}/{len(rows)}] No={out['No']} {out['研究室名']} status={out['URL取得ステータス']} confirm={out['問い生成確認状態']}")
        if args.sleep:
            time.sleep(args.sleep)

    write_csv(Path(args.output), OUTPUT_COLUMNS, outputs)
    write_csv(Path(args.report), WARNING_COLUMNS, logs)
    warning_rows = [log for log in logs if log["warning"]]
    write_csv(Path(args.warnings), WARNING_COLUMNS, warning_rows)

    success = sum(1 for r in outputs if r["URL取得ステータス"] == "success")
    unverified = sum(1 for r in outputs if r["問い生成確認状態"] == "unverified")
    related = sum(1 for r in outputs if r["関係教員_追加"])
    print(f"done rows={len(outputs)} success={success} unverified={unverified} related_teachers={related}")
    print(args.output)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
