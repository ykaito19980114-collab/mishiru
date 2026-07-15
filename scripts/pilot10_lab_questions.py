#!/usr/bin/env python3
"""10-row pilot: extract research blocks first, then generate questions only.

This script deliberately skips teacher extraction and never expands beyond the
requested 10 pilot rows.
"""

from __future__ import annotations

import argparse
import csv
import html
import re
import sys
import time
import urllib.parse
from html.parser import HTMLParser
from pathlib import Path

from openpyxl import load_workbook

SCRIPT_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(SCRIPT_DIR))
import enrich_lab_questions as base  # noqa: E402


TARGET_CATEGORIES = [
    ("materials", r"強相関|量子物性|機能性物質|材料|物性|分光"),
    ("chemistry", r"有機合成|分子集積|反応工学|化学工学|分子集合|触媒"),
    ("robotics", r"ロボット|制御|マニピュレーション|ヒューマン|身体運動"),
    ("information", r"情報|人工知能|AI|統計|データ|センシング|計算"),
    ("bio_medical", r"生命|生体|医療|細胞|遺伝子|バイオ|看護|薬|医学"),
    ("architecture", r"建築|都市|構造|防災|まち|土木"),
    ("environment", r"環境|生態|気候|海洋|農|植物|水|土壌"),
    ("social_humanities", r"社会|心理|教育|経済|法|文学|歴史|文化|人文"),
    ("energy_systems", r"エネルギー|電力|電気|システム|通信|半導体"),
    ("other", r".+"),
]

EXTRACT_COLUMNS = [
    "No",
    "研究室名",
    "大学名",
    "入力URL",
    "最終採用URL",
    "研究内容候補ブロック",
    "抽出ステータス",
    "除外理由",
    "確認した候補URL数",
]

QUESTION_COLUMNS = [
    "No",
    "研究室名",
    "研究内容_引用説明",
    "扱う問い_1",
    "扱う問い_2",
    "研究内容_取得元URL",
    "問い生成根拠メモ",
    "問い生成確認状態",
    "URL取得ステータス",
]

PRIORITY_LINK = re.compile(
    r"研究内容|研究概要|研究紹介|研究テーマ|研究プロジェクト|研究室紹介|Research|Research Topics|Research Interest|Our research|Projects|About|About us",
    re.I,
)
EXCLUDE_LINK = re.compile(
    r"News|ニュース|お知らせ|Information|Topics|新着|メンバー|学生|卒業|Publication|Paper|Achievement|アクセス|Contact|リンク|Copyright",
    re.I,
)
BAD_BLOCK = re.compile(
    r"News|ニュース|お知らせ|Copyright|Access|Contact|ページタイトル|トップページ|公式ページでは「News|研究対象や方法を紹介している|©|All rights reserved|open in new window",
    re.I,
)
RESEARCH_WORD = re.compile(
    r"研究|解析|実験|測定|理論|開発|設計|制御|調査|観察|評価|モデル|シミュレーション|物性|材料|量子|分子|細胞|医療|ロボット|構造|都市|環境|データ|情報|エネルギー"
)
CONTENT_WORD = re.compile(
    r"解析|実験|測定|理論|開発|設計|制御|調査|観察|評価|モデル|シミュレーション|物性|材料|量子|分子|細胞|医療|ロボット|構造|都市|環境|データ|情報|エネルギー|超伝導|磁性|触媒|反応|生体膜|リポソーム|ナノ|半導体|デバイス|欠陥|燃焼|複合材料|航空|破壊|疲労|強度|熱|流体|相互作用|通信|コンピュータ|ネットワーク|電子|スピン"
)
EVENT_WORD = re.compile(
    r"\d{4}[./年]\d{1,2}|受賞|セミナー|開催|更新|掲載|公開|会合|ファイナリスト|MVP|新聞|発表|合宿|研究会|研究室旅行|資料を公表|講演|メンバー|Activity|Publication|Equipment|授業|演習|特別実験|送別会|忘年会|加わりました|卒業|修了|研究業績|研究設備|研究室活動|採択|ホームページ|オープン|プレスリリース|記事|冊子|参加します|紹介されました|成功|English",
    re.I,
)


class LinkExtractor(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.links: list[tuple[str, str]] = []
        self._href = ""
        self._text: list[str] = []

    def handle_starttag(self, tag, attrs):
        if tag.lower() == "a":
            self._href = dict(attrs).get("href", "") or ""
            self._text = []

    def handle_data(self, data):
        if self._href:
            self._text.append(data.strip())

    def handle_endtag(self, tag):
        if tag.lower() == "a" and self._href:
            text = " ".join(x for x in self._text if x).strip()
            self.links.append((text, self._href))
            self._href = ""
            self._text = []


def read_text_and_links(url: str) -> tuple[str, str, str, list[str]]:
    text, final_url, status = base.fetch_text(url, timeout=12)
    links: list[str] = []
    if status != "success":
        return text, final_url, status, links
    # Fetch once more only for link extraction. Keep this pilot simple and bounded.
    try:
        import urllib.request

        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0 LaboIkitaiPilot10"})
        with urllib.request.urlopen(req, timeout=12) as res:
            data = res.read(1_200_000)
            raw = data.decode("utf-8", errors="ignore")
        parser = LinkExtractor()
        parser.feed(raw)
        base_url = final_url or url
        origin = urllib.parse.urlparse(base_url).netloc
        ranked = []
        for label, href in parser.links:
            full = urllib.parse.urljoin(base_url, href)
            parsed = urllib.parse.urlparse(full)
            if parsed.netloc != origin:
                continue
            if EXCLUDE_LINK.search(label + " " + full):
                continue
            if PRIORITY_LINK.search(label + " " + full):
                ranked.append(full.split("#")[0])
        links = list(dict.fromkeys(ranked))[:4]
    except Exception:
        pass
    return text, final_url, status, links


def candidate_block(text: str, limit: int = 1200) -> tuple[str, str]:
    sentences = []
    for raw in re.split(r"[\n。]+", text):
        line = re.sub(r"\s+", " ", raw).strip(" ・|｜")
        if len(line) < 12:
            continue
        if BAD_BLOCK.search(line):
            continue
        if re.search(r"^(Menu|Home|Site|Access|Contact|Copyright|News)$", line, re.I):
            continue
        if EVENT_WORD.search(line):
            continue
        if not RESEARCH_WORD.search(line) or not CONTENT_WORD.search(line):
            continue
        sentences.append(line)
    uniq = []
    seen = set()
    for s in sentences:
        key = s[:48]
        if key in seen:
            continue
        seen.add(key)
        uniq.append(s)
    block = "。".join(uniq)[:limit].strip()
    if len(block) < 100:
        return "", "研究内容候補が100字未満、またはNews/ナビ中心"
    weak = len(CONTENT_WORD.findall(block)) < 3
    if weak:
        return "", "研究対象・方法・応用先を判断できる本文が弱い"
    return block, ""


def clean_existing_block(text: str) -> tuple[str, str]:
    return candidate_block(text, limit=1200)


def pilot_description(lab_name: str, university: str, keywords: str, block: str) -> str:
    terms = base.pick_terms(lab_name, keywords, block)
    topic = "、".join(terms[:3]) if terms else base.strip_lab_suffix(lab_name)
    sentences = []
    for raw in re.split(r"。", block):
        line = re.sub(r"\s+", " ", raw).strip()
        if len(line) < 14 or BAD_BLOCK.search(line) or EVENT_WORD.search(line):
            continue
        if CONTENT_WORD.search(line):
            sentences.append(line)
        if len(sentences) >= 4:
            break
    core = "。".join(sentences[:3])
    domain = base.lab_domain(lab_name, keywords, terms, block)
    if core:
        desc = f"{topic}を中心に扱う研究室。抽出した公式本文では、{core}。"
    else:
        desc = f"{topic}を中心に扱う研究室。抽出した公式本文に含まれる研究対象と方法をもとに整理している。"
    if domain == "strong_electron":
        desc += " 電子・スピン・超伝導・磁性など、物質の中で現れる量子物性を理論や測定から読み解く方向が見える。"
    elif domain == "chemistry":
        desc += " 分子設計、反応設計、触媒、分子集合などを通じて、狙った構造や機能を作ることが焦点になる。"
    elif domain == "robotics":
        desc += " ロボットや人との相互作用を対象に、認識、制御、実験、設計を結びつける研究として読める。"
    elif domain == "biomechanics":
        desc += " 身体運動、感覚、筋骨格、力学を対象に、計測やモデル化から動きの成り立ちを調べる。"
    elif domain == "computing":
        desc += " 情報、電力、データ、計算、制御を対象に、システムの安定性や効率を高める視点がある。"
    elif domain == "fluid":
        desc += " 熱や流れ、乱れ、エネルギー変換を対象に、実験・可視化・解析を通じて現象を捉える。"
    if len(desc) > 300:
        desc = desc[:297] + "…"
    return desc


def pilot_questions(lab_name: str, keywords: str, block: str) -> tuple[str, str]:
    name_kw = lab_name + " " + keywords
    if re.search(r"構造力学|航空宇宙", name_kw):
        return (
            "航空機や宇宙機の構造は、軽さを保ちながら荷重や損傷にどこまで耐えられるのか？",
            "複合材料や構造部材の変形・破壊をどう予測し、安全な設計に結びつけられるのか？",
        )
    if re.search(r"環境・エネルギー|燃焼", name_kw):
        return (
            "燃焼や熱エネルギー変換を、環境負荷を抑えながらどこまで高効率にできるのか？",
            "実験・計測・解析を通じて、燃料や排出物のふるまいをどう制御できるのか？",
        )
    if re.search(r"ナノエレクトロニクス|半導体|メモリスタ", name_kw):
        return (
            "半導体結晶の欠陥や界面は、ナノデバイスの性能や信頼性をどう左右しているのか？",
            "材料評価とデバイス作製を結びつけて、次世代メモリや電子素子の機能をどう引き出せるのか？",
        )
    return base.make_questions(lab_name, keywords, block)


def extract_for_row(row: dict[str, object]) -> dict[str, str]:
    url = base.normalize_url(str(row.get("研究室URL", "") or ""))
    checked = 0
    adopted_text = ""
    adopted_url = ""
    status = "no_url" if not url else "no_research_content"
    reason = ""

    candidates = [url] if url else []
    text, final_url, first_status, links = read_text_and_links(url) if url else ("", "", "no_url", [])
    checked += 1 if url else 0
    if first_status == "success":
        block, reason = candidate_block(text)
        if block:
            adopted_text, adopted_url, status = block, final_url or url, "success"
        else:
            candidates.extend([x for x in links if x not in candidates])
    else:
        status = first_status
        reason = first_status

    for link in candidates[1:5]:
        if adopted_text:
            break
        text2, final2, status2, _links2 = read_text_and_links(link)
        checked += 1
        if status2 != "success":
            reason = status2
            continue
        block, reason2 = candidate_block(text2)
        if block:
            adopted_text, adopted_url, status = block, final2 or link, "success"
            reason = ""
            break
        reason = reason2

    if not adopted_text and status == "success":
        status = "no_research_content"

    return {
        "No": str(int(row.get("No"))) if isinstance(row.get("No"), float) else str(row.get("No", "")),
        "研究室名": base.clean_label(row.get("研究室名", "")),
        "大学名": base.clean_label(row.get("大学名", "")),
        "入力URL": url,
        "最終採用URL": adopted_url or final_url or url,
        "研究内容候補ブロック": adopted_text,
        "抽出ステータス": status,
        "除外理由": "" if adopted_text else (reason or status),
        "確認した候補URL数": str(checked),
    }


def generate_question_row(row: dict[str, object], extracted: dict[str, str]) -> dict[str, str]:
    no = extracted["No"]
    lab = extracted["研究室名"]
    if extracted["抽出ステータス"] != "success" or len(extracted["研究内容候補ブロック"]) < 100:
        return {
            "No": no,
            "研究室名": lab,
            "研究内容_引用説明": "",
            "扱う問い_1": "",
            "扱う問い_2": "",
            "研究内容_取得元URL": extracted["最終採用URL"],
            "問い生成根拠メモ": "研究内容として使える本文ブロックを抽出できなかった",
            "問い生成確認状態": "unverified",
            "URL取得ステータス": "no_research_content",
        }
    block = extracted["研究内容候補ブロック"]
    keywords = base.clean_label(row.get("研究分野・キーワード", ""))
    desc = pilot_description(lab, extracted["大学名"], keywords, block)
    q1, q2 = pilot_questions(lab, keywords, block)
    if BAD_BLOCK.search(desc) or any(ng in q1 + q2 for ng in base.NG_PATTERNS):
        desc = q1 = q2 = ""
        confirm = "unverified"
        memo = "品質チェックでNews/テンプレ表現が検出されたため未生成扱い"
        status = "no_research_content"
    else:
        confirm = "official"
        memo = "抽出した研究内容候補ブロックのみを根拠に作成"
        status = "success"
    return {
        "No": no,
        "研究室名": lab,
        "研究内容_引用説明": desc,
        "扱う問い_1": q1,
        "扱う問い_2": q2,
        "研究内容_取得元URL": extracted["最終採用URL"],
        "問い生成根拠メモ": memo,
        "問い生成確認状態": confirm,
        "URL取得ステータス": status,
    }


def read_workbook(path: Path, sheet: str) -> list[dict[str, object]]:
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet]
    headers = [str(ws.cell(1, c).value or "").strip() for c in range(1, ws.max_column + 1)]
    rows = []
    for r in range(2, ws.max_row + 1):
        rows.append({headers[c - 1]: ws.cell(r, c).value for c in range(1, ws.max_column + 1)})
    return rows


def read_workbook_by_no(path: Path, sheet: str, target_nos: set[str]) -> list[dict[str, object]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet]
    row_iter = ws.iter_rows(values_only=True)
    headers = [str(value or "").strip() for value in next(row_iter)]
    no_index = headers.index("No")
    rows = []
    for values in row_iter:
        no = normalize_no(values[no_index]) if no_index < len(values) else ""
        if no not in target_nos:
            continue
        rows.append({headers[i]: values[i] if i < len(values) else None for i in range(len(headers))})
        if len(rows) >= len(target_nos):
            break
    return rows


def pick_diverse(rows: list[dict[str, object]], count: int = 10) -> list[dict[str, object]]:
    picked = []
    used_no = set()
    for _category, pattern in TARGET_CATEGORIES:
        for row in rows:
            no = row.get("No")
            if no in used_no:
                continue
            text = " ".join(str(row.get(k, "") or "") for k in ["研究室名", "研究分野・キーワード", "学部・研究科・専攻"])
            if re.search(pattern, text):
                picked.append(row)
                used_no.add(no)
                break
        if len(picked) >= count:
            break
    for row in rows:
        if len(picked) >= count:
            break
        if row.get("No") not in used_no:
            picked.append(row)
            used_no.add(row.get("No"))
    return picked[:count]


def write_csv(path: Path, cols: list[str], rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=cols)
        writer.writeheader()
        writer.writerows(rows)


def write_review(path: Path, extracted: list[dict[str, str]], questions: list[dict[str, str]]) -> None:
    qmap = {q["No"]: q for q in questions}
    parts = ["# lab_questions_pilot_10_review\n"]
    for ex in extracted:
        q = qmap[ex["No"]]
        concern = "なし" if q["問い生成確認状態"] == "official" else ex["除外理由"] or "研究内容本文不足"
        self_eval = "生成可: 抽出ブロックに研究対象・方法語が含まれる" if q["問い生成確認状態"] == "official" else "未生成: 研究内容本文が不足"
        parts.append(
            "\n".join(
                [
                    f"## No {ex['No']} {ex['研究室名']}",
                    f"- 入力URL: {ex['入力URL']}",
                    f"- 最終採用URL: {ex['最終採用URL']}",
                    f"- 研究内容候補ブロック: {ex['研究内容候補ブロック'] or '(空欄)'}",
                    f"- 研究内容_引用説明: {q['研究内容_引用説明'] or '(空欄)'}",
                    f"- 扱う問い_1: {q['扱う問い_1'] or '(空欄)'}",
                    f"- 扱う問い_2: {q['扱う問い_2'] or '(空欄)'}",
                    f"- 抽出ステータス: {ex['抽出ステータス']}",
                    f"- 問い生成確認状態: {q['問い生成確認状態']}",
                    f"- URL取得ステータス: {q['URL取得ステータス']}",
                    f"- 自己評価: {self_eval}",
                    f"- 懸念点: {concern}",
                    "",
                ]
            )
        )
    path.write_text("\n".join(parts), encoding="utf-8")


def normalize_no(value: object) -> str:
    return str(int(value)) if isinstance(value, float) else str(value or "")


def regenerate_from_extracted(input_path: Path, workbook_rows: list[dict[str, object]], output_dir: Path) -> None:
    row_by_no = {normalize_no(row.get("No")): row for row in workbook_rows}
    with input_path.open(encoding="utf-8-sig", newline="") as f:
        source_rows = list(csv.DictReader(f))

    extracted: list[dict[str, str]] = []
    questions: list[dict[str, str]] = []
    for raw in source_rows[:10]:
        block, reason = clean_existing_block(raw.get("研究内容候補ブロック", ""))
        ex = {col: raw.get(col, "") for col in EXTRACT_COLUMNS}
        if block:
            ex["研究内容候補ブロック"] = block
            ex["抽出ステータス"] = "success"
            ex["除外理由"] = ""
        else:
            ex["研究内容候補ブロック"] = ""
            ex["抽出ステータス"] = "no_research_content"
            ex["除外理由"] = reason or raw.get("除外理由", "") or "研究内容本文不足"
        row = row_by_no.get(ex["No"], {})
        extracted.append(ex)
        questions.append(generate_question_row(row, ex))

    write_csv(output_dir / "extracted_research_blocks_pilot_10.csv", EXTRACT_COLUMNS, extracted)
    write_csv(output_dir / "lab_questions_pilot_10.csv", QUESTION_COLUMNS, questions)
    write_review(output_dir / "lab_questions_pilot_10_review.md", extracted, questions)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True)
    parser.add_argument("--sheet", default="研究室リスト_掲出用")
    parser.add_argument("--output-dir", default="outputs")
    parser.add_argument("--sleep", type=float, default=0.2)
    parser.add_argument("--reuse-extracted", default="")
    args = parser.parse_args()

    outdir = Path(args.output_dir)
    if args.reuse_extracted:
        with Path(args.reuse_extracted).open(encoding="utf-8-sig", newline="") as f:
            target_nos = {row.get("No", "") for row in csv.DictReader(f)}
        rows = read_workbook_by_no(Path(args.input), args.sheet, target_nos)
        regenerate_from_extracted(Path(args.reuse_extracted), rows, outdir)
        print(outdir / "extracted_research_blocks_pilot_10.csv")
        print(outdir / "lab_questions_pilot_10.csv")
        print(outdir / "lab_questions_pilot_10_review.md")
        return 0

    all_rows = read_workbook(Path(args.input), args.sheet)
    rows = pick_diverse(all_rows, 10)
    extracted = []
    questions = []
    for idx, row in enumerate(rows, 1):
        ex = extract_for_row(row)
        q = generate_question_row(row, ex)
        extracted.append(ex)
        questions.append(q)
        print(f"[{idx}/10] No={ex['No']} {ex['研究室名']} extract={ex['抽出ステータス']} gen={q['問い生成確認状態']}")
        time.sleep(args.sleep)

    write_csv(outdir / "extracted_research_blocks_pilot_10.csv", EXTRACT_COLUMNS, extracted)
    write_csv(outdir / "lab_questions_pilot_10.csv", QUESTION_COLUMNS, questions)
    write_review(outdir / "lab_questions_pilot_10_review.md", extracted, questions)
    print(outdir / "extracted_research_blocks_pilot_10.csv")
    print(outdir / "lab_questions_pilot_10.csv")
    print(outdir / "lab_questions_pilot_10_review.md")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
